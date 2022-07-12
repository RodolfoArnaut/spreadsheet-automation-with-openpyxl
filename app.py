import openpyxl

#criando a planilha
 
book = openpyxl.Workbook()

#visualizando páginas existentes

print (book.sheetnames)

#Criando a página 

book.create_sheet('Produtos')

#selecionando uma página 

produtos_page  =  book['Produtos']
produtos_page.append (['Produto', 'Quantidade', 'Valor'])
produtos_page.append ([ "Computador 1", '5', '3100'])
produtos_page.append (["Computador 2", '5', '3500'])
produtos_page.append (["Computador 3", '2', '6400'])
produtos_page.append (["Computador 4", '7', '2100'])
produtos_page.append (["Computador 5", '9', '4000'])

#salvar a planilha

book.save('Planilha Produtos.xlsx')